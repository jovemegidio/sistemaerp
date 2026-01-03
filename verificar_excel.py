# Verificacao do arquivo Excel gerado
import openpyxl

wb = openpyxl.load_workbook('teste_ordem_gerada.xlsx')
ws = wb['VENDAS_PCP']

print('========================================')
print('VERIFICACAO DO ARQUIVO EXCEL GERADO')
print('========================================')
print()

# Verificar linha 14 (titulo)
print('LINHA 14 (Dados para Cobranca - TITULO):')
print('   A14:', repr(ws['A14'].value))
print('   C14:', repr(ws['C14'].value), '<- DEVE SER VAZIO!')
print('   G14:', repr(ws['G14'].value), '<- DEVE SER VAZIO!')
print()

# Verificar linha 15 (dados)
print('LINHA 15 (CPF/CNPJ e Email):')
print('   A15:', repr(ws['A15'].value))
print('   C15:', repr(ws['C15'].value), '<- DEVE TER CPF/CNPJ DO CLIENTE')
print('   G15:', repr(ws['G15'].value), '<- DEVE TER EMAIL')
print()

# Verificar dados do cliente
print('DADOS DO CLIENTE:')
print('   C7 (Cliente):', ws['C7'].value)
print('   C8 (Contato):', ws['C8'].value)
print('   H8 (Telefone):', ws['H8'].value)
print('   C9 (Email):', ws['C9'].value)
print()

# Verificar produtos
print('PRODUTOS:')
for i in range(18, 21):
    cod = ws['B' + str(i)].value
    qtd = ws['H' + str(i)].value
    preco = ws['I' + str(i)].value
    total = ws['J' + str(i)].value
    if cod:
        print('   Linha', i, ':', cod, '| Qtd:', qtd, '| Preco:', preco, '| Total:', total)
print()

# Resumo
print('========================================')
print('RESULTADO:')
c14_vazio = ws['C14'].value is None or ws['C14'].value == ''
g14_vazio = ws['G14'].value is None or ws['G14'].value == ''
c15_preenchido = ws['C15'].value is not None and ws['C15'].value != ''
g15_preenchido = ws['G15'].value is not None and ws['G15'].value != ''

if c14_vazio and g14_vazio:
    print('   [OK] Linha 14 vazia corretamente (apenas titulo)')
else:
    print('   [ERRO] Linha 14 tem dados indevidos!')

if c15_preenchido:
    print('   [OK] C15 tem CPF/CNPJ:', ws['C15'].value)
else:
    print('   [ERRO] C15 esta vazio!')

if g15_preenchido:
    print('   [OK] G15 tem Email:', ws['G15'].value)
else:
    print('   [ERRO] G15 esta vazio!')

print('========================================')
