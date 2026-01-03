import openpyxl
import json

# Carregar o arquivo Excel modelo
wb = openpyxl.load_workbook('modules/PCP/Ordem de Produção Aluforce - Copia.xlsx', data_only=False)
ws = wb.active

print('='*80)
print('MAPEAMENTO COMPLETO DE FÓRMULAS E CÉLULAS')
print('='*80)

# Mapear todas as células com fórmulas
formulas = {}
valores_estaticos = {}
merged_cells = []

# Células mescladas
for merged in ws.merged_cells.ranges:
    merged_cells.append(str(merged))

print(f'\nCELULAS MESCLADAS ({len(merged_cells)}):')
for i, merged in enumerate(merged_cells[:20], 1):
    print(f'{i}. {merged}')

print('\n' + '='*80)
print('CABEÇALHO E DADOS DO PEDIDO (Linhas 4-15)')
print('='*80)

for row in range(4, 16):
    for col in range(1, 12):
        cell = ws.cell(row, col)
        col_letter = openpyxl.utils.get_column_letter(col)
        cell_ref = f'{col_letter}{row}'
        
        if cell.value:
            if isinstance(cell.value, str) and cell.value.startswith('='):
                formulas[cell_ref] = cell.value
                print(f'\n[FORMULA] {cell_ref}: {cell.value}')
            elif not isinstance(cell.value, str) or len(str(cell.value)) < 30:
                valores_estaticos[cell_ref] = cell.value
                print(f'[VALOR] {cell_ref}: {cell.value}')

print('\n' + '='*80)
print('TABELA DE PRODUTOS (Linhas 12-56)')
print('='*80)

for row in range(12, 57):
    for col in [1, 2, 3, 4, 5, 6, 7, 8, 9, 10]:
        cell = ws.cell(row, col)
        col_letter = openpyxl.utils.get_column_letter(col)
        cell_ref = f'{col_letter}{row}'
        
        if cell.value:
            if isinstance(cell.value, str) and cell.value.startswith('='):
                formulas[cell_ref] = cell.value
                print(f'\n[FORMULA PRODUTO] {cell_ref}: {cell.value}')

print('\n' + '='*80)
print('BANCO DE DADOS DE PRODUTOS (Linhas 18-50, Colunas N-P)')
print('='*80)

produtos_db = []
for row in range(18, 51):
    n_val = ws.cell(row, 14).value  # Coluna N
    o_val = ws.cell(row, 15).value  # Coluna O
    p_val = ws.cell(row, 16).value  # Coluna P
    
    if n_val or o_val or p_val:
        produto = {
            'linha': row,
            'codigo': n_val,
            'descricao': o_val,
            'cor': p_val
        }
        produtos_db.append(produto)
        print(f'Linha {row}: {n_val} | {o_val} | {p_val}')

print('\n' + '='*80)
print('RESUMO DAS FÓRMULAS ENCONTRADAS')
print('='*80)

# Agrupar fórmulas por tipo
vlookup_formulas = {k: v for k, v in formulas.items() if 'VLOOKUP' in v.upper()}
vendas_refs = {k: v for k, v in formulas.items() if 'VENDAS_PCP' in v.upper()}
outras_formulas = {k: v for k, v in formulas.items() if 'VLOOKUP' not in v.upper() and 'VENDAS_PCP' not in v.upper()}

print(f'\n[TOTAL] Formulas: {len(formulas)}')
print(f'   - Referencias VENDAS_PCP: {len(vendas_refs)}')
print(f'   - VLOOKUP: {len(vlookup_formulas)}')
print(f'   - Outras: {len(outras_formulas)}')

print('\n[VENDAS_PCP] Referencias a planilha externa:')
for cell_ref, formula in list(vendas_refs.items())[:15]:
    print(f'   {cell_ref}: {formula}')

print('\n[VLOOKUP] Formulas de busca:')
for cell_ref, formula in list(vlookup_formulas.items())[:15]:
    print(f'   {cell_ref}: {formula}')

print('\n' + '='*80)
print('ESTRUTURA JSON PARA IMPLEMENTAÇÃO')
print('='*80)

estrutura = {
    'celulas_mescladas': merged_cells[:20],
    'dados_pedido': {
        'linha_4': {
            'C4': 'Orçamento (VENDAS_PCP!C4)',
            'E4': 'Revisão (VENDAS_PCP!E4)',
            'G4': 'Pedido (VENDAS_PCP!G4)',
            'J4': 'Data Liberação (VENDAS_PCP!J4)'
        },
        'linha_6': {
            'C6': 'Vendedor (VENDAS_PCP!C6)',
            'H6': 'Prazo Entrega (VENDAS_PCP!H6)'
        },
        'linha_7': {
            'C7': 'Cliente (VENDAS_PCP!C7)'
        },
        'linha_8': {
            'C8': 'Contato (VENDAS_PCP!C8)',
            'H8': 'Fone (VENDAS_PCP!H8)'
        },
        'linha_9': {
            'C9': 'Email (VENDAS_PCP!C9)',
            'J9': 'Frete (VENDAS_PCP!J9)'
        }
    },
    'produtos': {
        'coluna_B': 'Código (=VENDAS_PCP!B18)',
        'coluna_C': 'Descrição (=VLOOKUP(B13,N18:O175,2,0))',
        'coluna_F': 'Cod Cores (=VLOOKUP(B13,N18:P184,3,0))',
        'coluna_H': 'Embalagem (=VENDAS_PCP!F18)',
        'coluna_I': 'Lances (=VENDAS_PCP!G18)',
        'coluna_J': 'Quantidade (=VENDAS_PCP!H18)'
    },
    'banco_produtos': [p for p in produtos_db[:10]],
    'total_formulas': len(formulas),
    'total_produtos_db': len(produtos_db)
}

print(json.dumps(estrutura, indent=2, ensure_ascii=False))

wb.close()

print('\n' + '='*80)
print('ANÁLISE CONCLUÍDA')
print('='*80)
