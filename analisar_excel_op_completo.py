"""
Análise completa do template Excel de Ordem de Produção
"""
import openpyxl
from openpyxl.utils import get_column_letter
import json

# Caminho do arquivo
arquivo = r"c:\Users\egidio\Documents\Sistema - ALUFORCE - V.2\modules\PCP\Ordem de Produção Aluforce - Copia.xlsx"

print("=" * 80)
print("ANÁLISE COMPLETA DO TEMPLATE EXCEL - ORDEM DE PRODUÇÃO")
print("=" * 80)

try:
    wb = openpyxl.load_workbook(arquivo, data_only=False)
    print(f"\nPlanilhas encontradas: {wb.sheetnames}")
    
    # Análise de cada planilha
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\n{'='*80}")
        print(f"PLANILHA: {sheet_name}")
        print(f"{'='*80}")
        print(f"Dimensões: {ws.dimensions}")
        print(f"Linhas usadas: {ws.max_row}, Colunas usadas: {ws.max_column}")
        
        # Analisar todas as células com conteúdo
        print(f"\n--- TODAS AS CÉLULAS COM CONTEÚDO ---")
        
        for row in range(1, min(ws.max_row + 1, 50)):  # Primeiras 50 linhas
            for col in range(1, min(ws.max_column + 1, 15)):  # Primeiras 15 colunas
                cell = ws.cell(row=row, column=col)
                if cell.value is not None or cell.fill.start_color.rgb != '00000000':
                    col_letter = get_column_letter(col)
                    cell_ref = f"{col_letter}{row}"
                    
                    # Verificar se é fórmula
                    is_formula = str(cell.value).startswith('=') if cell.value else False
                    
                    # Verificar cor de fundo
                    try:
                        fill_color = cell.fill.start_color.rgb if cell.fill.start_color.rgb else "None"
                    except:
                        fill_color = "None"
                    
                    if cell.value is not None:
                        value_display = str(cell.value)[:100] if cell.value else ""
                        tipo = "FÓRMULA" if is_formula else "VALOR"
                        print(f"  [{cell_ref:6}] ({tipo:8}) = {value_display}")
        
        # Se for VENDAS_PCP, análise detalhada
        if 'VENDAS' in sheet_name.upper() or 'PCP' in sheet_name.upper():
            print(f"\n--- ANÁLISE DETALHADA VENDAS_PCP ---")
            
            print("\n>> CABEÇALHO (Linhas 1-17):")
            for row in range(1, 18):
                row_data = []
                for col in range(1, 12):
                    cell = ws.cell(row=row, column=col)
                    if cell.value:
                        col_letter = get_column_letter(col)
                        row_data.append(f"{col_letter}:{cell.value}")
                if row_data:
                    print(f"  Linha {row:2}: {' | '.join(row_data)}")
            
            print("\n>> ÁREA DE PRODUTOS (Linhas 18-35):")
            print("  Colunas: ", end="")
            for col in range(1, 12):
                col_letter = get_column_letter(col)
                header_cell = ws.cell(row=17, column=col)
                if header_cell.value:
                    print(f"{col_letter}={header_cell.value} | ", end="")
            print()
            
            for row in range(18, 36):
                row_data = []
                for col in range(1, 12):
                    cell = ws.cell(row=row, column=col)
                    col_letter = get_column_letter(col)
                    if cell.value:
                        is_formula = str(cell.value).startswith('=') if cell.value else False
                        tipo = "F" if is_formula else "V"
                        row_data.append(f"{col_letter}[{tipo}]:{str(cell.value)[:30]}")
                    else:
                        row_data.append(f"{col_letter}:___")
                print(f"  Linha {row}: {' | '.join(row_data)}")
        
        # Se for PRODUÇÃO, análise detalhada
        if 'PRODU' in sheet_name.upper():
            print(f"\n--- ANÁLISE DETALHADA PRODUÇÃO ---")
            
            print("\n>> CABEÇALHO:")
            for row in range(1, 20):
                row_data = []
                for col in range(1, 15):
                    cell = ws.cell(row=row, column=col)
                    if cell.value:
                        col_letter = get_column_letter(col)
                        is_formula = str(cell.value).startswith('=') if cell.value else False
                        tipo = "F" if is_formula else "V"
                        row_data.append(f"{col_letter}[{tipo}]:{str(cell.value)[:40]}")
                if row_data:
                    print(f"  Linha {row:2}: {' | '.join(row_data)}")
            
            # Buscar referências a outras planilhas
            print("\n>> REFERÊNCIAS A OUTRAS PLANILHAS:")
            for row in range(1, ws.max_row + 1):
                for col in range(1, ws.max_column + 1):
                    cell = ws.cell(row=row, column=col)
                    if cell.value and '!' in str(cell.value):
                        col_letter = get_column_letter(col)
                        print(f"  [{col_letter}{row}]: {cell.value}")

    # Buscar campos específicos em todas as planilhas
    print(f"\n{'='*80}")
    print("BUSCA POR CAMPOS ESPECÍFICOS")
    print(f"{'='*80}")
    
    campos_busca = ['COR', 'CORES', 'CÓDIGO', 'CODIGO', 'COD', 'PT', 'CZ', 'NU', 
                    'PRETO', 'CINZA', 'CLIENTE', 'PEDIDO', 'DATA', 'DESCRIÇÃO',
                    'QTD', 'QUANTIDADE', 'LINHA', 'PRODUTO', 'ITEM', 'MODELO']
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\n>> Planilha: {sheet_name}")
        
        for row in range(1, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                valor = str(cell.value).upper() if cell.value else ""
                
                for campo in campos_busca:
                    if campo in valor:
                        col_letter = get_column_letter(col)
                        is_formula = str(cell.value).startswith('=') if cell.value else False
                        tipo = "FÓRMULA" if is_formula else "VALOR"
                        print(f"  [{col_letter}{row}] ({tipo}): {cell.value}")
                        break

    wb.close()
    
except Exception as e:
    print(f"ERRO: {e}")
    import traceback
    traceback.print_exc()
