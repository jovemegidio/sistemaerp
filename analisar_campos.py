# Analise detalhada do Excel gerado
import openpyxl

wb = openpyxl.load_workbook('teste_ordem_gerada.xlsx')

# Verificar VENDAS_PCP
ws = wb['VENDAS_PCP']
print('=== ABA VENDAS_PCP ===')
print()
print('TRANSPORTADORA (Linhas 11-13):')
print('   A11:', ws['A11'].value)
print('   C12 (Nome):', ws['C12'].value)
print('   H12 (Fone):', ws['H12'].value)
print('   C13 (CEP):', ws['C13'].value)
print('   F13 (Endereco):', ws['F13'].value)
print()

print('PRODUTOS (Linhas 18-20):')
for i in range(18, 21):
    item = ws['A' + str(i)].value
    codigo = ws['B' + str(i)].value
    descricao = ws['C' + str(i)].value
    embalagem = ws['F' + str(i)].value
    lances = ws['G' + str(i)].value
    qtd = ws['H' + str(i)].value
    preco = ws['I' + str(i)].value
    
    print('   Linha', i, ':')
    print('      A (Item):', item)
    print('      B (Codigo):', codigo)
    print('      C (Descricao):', descricao)
    print('      F (Embalagem):', embalagem)
    print('      G (Lances):', lances)
    print('      H (Qtd):', qtd)
    print('      I (Preco):', preco)
    print()

# Verificar catalogo de produtos (colunas N:O)
print('CATALOGO DE PRODUTOS (N:O):')
for i in range(18, 25):
    cod = ws['N' + str(i)].value
    desc = ws['O' + str(i)].value
    if cod:
        print('   N' + str(i) + ':', cod, '->', desc)

# Verificar PRODUCAO
print()
print('=== ABA PRODUCAO ===')
if 'PRODUÇÃO' in wb.sheetnames:
    ws2 = wb['PRODUÇÃO']
    print('Linhas de produtos: 13, 16, 19')
    for linha in [13, 16, 19]:
        codigo = ws2['B' + str(linha)].value
        produto = ws2['C' + str(linha)].value
        cod_cores = ws2['F' + str(linha)].value
        emb = ws2['I' + str(linha)].value if ws2['I' + str(linha)].value else ''
        lance = ws2['J' + str(linha)].value if ws2['J' + str(linha)].value else ''
        
        print('   Linha', linha, ':')
        print('      B (Codigo):', codigo)
        print('      C (Produto):', produto)
        print('      F (Cod Cores):', cod_cores)
        print('      I (Embalagem):', emb)
        print('      J (Lance):', lance)
        print()
elif 'PRODUCAO' in wb.sheetnames:
    ws2 = wb['PRODUCAO']
    print('Aba encontrada como PRODUCAO')
else:
    print('Aba de producao nao encontrada!')
    print('Abas disponiveis:', wb.sheetnames)
