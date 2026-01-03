"""
Script para analisar a tabela de origem do VLOOKUP (colunas N e O)
"""
import openpyxl
from openpyxl.utils import get_column_letter

# Caminho do arquivo
arquivo = r"c:\Users\egidio\Documents\Sistema - ALUFORCE - V.2\modules\PCP\Ordem de Produção Aluforce - Copia.xlsx"

print("=" * 80)
print("ANÁLISE DA TABELA DE ORIGEM DO VLOOKUP (Colunas N, O, P)")
print("=" * 80)

# Carregar workbook
wb = openpyxl.load_workbook(arquivo, data_only=False)

# Verificar colunas N, O, P na planilha VENDAS_PCP
print("\n1. TABELA DE PRODUTOS NA PLANILHA VENDAS_PCP (Colunas N, O, P):")
print("-" * 60)

ws_vendas = wb['VENDAS_PCP']

# Verificar o range N18:O198 que é referenciado no VLOOKUP
print("\n   Range N18:O198 (primeiras 30 linhas):")
print("   " + "-" * 50)

for row in range(16, 50):  # Começar um pouco antes para ver cabeçalhos
    n_val = ws_vendas.cell(row=row, column=14).value  # Coluna N
    o_val = ws_vendas.cell(row=row, column=15).value  # Coluna O
    p_val = ws_vendas.cell(row=row, column=16).value  # Coluna P
    q_val = ws_vendas.cell(row=row, column=17).value  # Coluna Q
    
    if n_val or o_val or p_val:
        print(f"   Linha {row}: N={n_val} | O={o_val} | P={p_val} | Q={q_val}")

# Contar total de registros na tabela de produtos
print("\n2. TOTAL DE REGISTROS NA TABELA DE PRODUTOS (Coluna N com dados):")
print("-" * 60)

count = 0
for row in range(18, 200):
    n_val = ws_vendas.cell(row=row, column=14).value
    if n_val:
        count += 1
        
print(f"   Total de produtos na coluna N: {count}")

# Verificar também na planilha PRODUÇÃO
print("\n3. TABELA DE PRODUTOS NA PLANILHA PRODUÇÃO (Colunas N, O, P):")
print("-" * 60)

ws_prod = wb['PRODUÇÃO']

print("\n   Range N18:P200 (primeiras 30 linhas):")
print("   " + "-" * 50)

for row in range(16, 50):
    n_val = ws_prod.cell(row=row, column=14).value  # Coluna N
    o_val = ws_prod.cell(row=row, column=15).value  # Coluna O
    p_val = ws_prod.cell(row=row, column=16).value  # Coluna P
    
    if n_val or o_val or p_val:
        print(f"   Linha {row}: N={n_val} | O={o_val} | P={p_val}")

# Verificar a coluna B (códigos dos produtos) que é a chave do VLOOKUP
print("\n4. COLUNA B (CÓDIGOS) NA PLANILHA VENDAS_PCP:")
print("-" * 60)

for row in range(17, 35):
    b_val = ws_vendas.cell(row=row, column=2).value  # Coluna B
    c_val = ws_vendas.cell(row=row, column=3).value  # Coluna C (resultado do VLOOKUP)
    
    if b_val:
        # Mostrar se C tem fórmula ou valor
        c_display = c_val
        if isinstance(c_val, str) and c_val.startswith('='):
            c_display = "[FÓRMULA]"
        print(f"   Linha {row}: B={b_val} | C={c_display}")

# Verificar se há mais colunas ocultas ou dados auxiliares
print("\n5. TODAS AS COLUNAS COM DADOS NA LINHA 17 (CABEÇALHOS):")
print("-" * 60)

for col in range(1, ws_vendas.max_column + 1):
    val = ws_vendas.cell(row=17, column=col).value
    if val:
        print(f"   {get_column_letter(col)}17: {val}")

print("\n6. ESTRUTURA COMPLETA - PRODUTOS EM VENDAS_PCP (N18:O até fim):")
print("-" * 60)

produtos_vendas = []
for row in range(18, 200):
    n_val = ws_vendas.cell(row=row, column=14).value
    o_val = ws_vendas.cell(row=row, column=15).value
    
    if n_val:
        produtos_vendas.append({'linha': row, 'codigo': n_val, 'nome': o_val})

print(f"   Total de produtos encontrados: {len(produtos_vendas)}")
print("\n   Lista completa:")
for p in produtos_vendas:
    print(f"      {p['linha']}: {p['codigo']} -> {p['nome']}")

print("\n" + "=" * 80)
print("FIM DA ANÁLISE")
print("=" * 80)

wb.close()
