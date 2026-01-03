"""
Análise detalhada do template Excel - Mapeamento completo
"""
import openpyxl
from openpyxl.utils import get_column_letter
import json

arquivo = r"c:\Users\egidio\Documents\Sistema - ALUFORCE - V.2\modules\PCP\Ordem de Produção Aluforce - Copia.xlsx"

wb = openpyxl.load_workbook(arquivo, data_only=False)

print("=" * 120)
print("MAPEAMENTO COMPLETO DO TEMPLATE EXCEL")
print("=" * 120)

# ===========================
# ANÁLISE VENDAS_PCP
# ===========================
ws = wb['VENDAS_PCP']

print("\n" + "=" * 120)
print("📋 PLANILHA VENDAS_PCP - ANÁLISE COMPLETA")
print("=" * 120)

print("\n" + "-" * 80)
print("1️⃣ CABEÇALHO (Informações do Pedido) - Linhas 1-15")
print("-" * 80)

cabecalho_vendas = [
    ("E1", "Cabeçalho/Logo", "VALOR", "Pedido + Dados empresa"),
    ("A3", "Marcador", "VALOR", "X"),
    ("C4", "Nº Orçamento", "VALOR", "352"),
    ("E4", "Revisão", "VALOR", "(vazio ou número)"),
    ("G4", "Nº Pedido", "VALOR", "202500083"),
    ("J4", "Data Liberação", "VALOR", "2025-08-19 (data)"),
    ("C6", "Vendedor", "VALOR", "Marcia Scarcella"),
    ("H6", "Prazo Entrega", "FÓRMULA", "=J4+30 (calcula automaticamente)"),
    ("C7", "Cliente", "VALOR", "CONSTRULAR"),
    ("C8", "Contato", "VALOR", "Rodrigo"),
    ("H8", "Telefone", "VALOR", "94984306216"),
    ("C9", "Email", "VALOR", "constrularcimento@gmail.com"),
    ("J9", "Tipo Frete", "VALOR", "FOB ou CIF"),
    ("C12", "Nome Transportadora", "VALOR", "(nome)"),
    ("H12", "Fone Transportadora", "FÓRMULA", "=H8 (copia telefone cliente)"),
    ("C13", "CEP", "VALOR", "68560-000"),
    ("F13", "Endereço Completo", "VALOR", "Av. Henrique Vita n°12..."),
    ("C15", "CPF/CNPJ", "VALOR", "36408556000169"),
    ("G15", "Email NFe", "FÓRMULA", "=C9 (copia email)"),
]

print(f"\n{'Célula':<8} | {'Descrição':<25} | {'Tipo':<10} | {'Exemplo/Valor'}")
print("-" * 100)
for cell, desc, tipo, exemplo in cabecalho_vendas:
    print(f"{cell:<8} | {desc:<25} | {tipo:<10} | {exemplo}")

print("\n" + "-" * 80)
print("2️⃣ CABEÇALHO DE PRODUTOS - Linha 17")
print("-" * 80)

colunas_produtos = [
    ("A17", "Item", "Número sequencial (1, 2, 3...)"),
    ("B17", "Cod.", "Código do produto (TRN10, DUN16, etc)"),
    ("C17", "Produto", "Descrição (fórmula VLOOKUP)"),
    ("D17", "-", "Não usado"),
    ("E17", "-", "Não usado"),
    ("F17", "Embalagem:", "Tipo (Bobina, Rolo, Lance)"),
    ("G17", "Lance(s)", "Formato lances (1x1000, 2x500)"),
    ("H17", "Qtd.", "Quantidade total"),
    ("I17", "V. Un. R$", "Valor unitário"),
    ("J17", "V. Total. R$", "Valor total (fórmula)"),
]

print(f"\n{'Célula':<8} | {'Cabeçalho':<15} | {'Descrição'}")
print("-" * 80)
for cell, header, desc in colunas_produtos:
    print(f"{cell:<8} | {header:<15} | {desc}")

print("\n" + "-" * 80)
print("3️⃣ ÁREA DE PRODUTOS - Linhas 18-32 (15 produtos máximo)")
print("-" * 80)

print("\n📌 ESTRUTURA DE CADA LINHA DE PRODUTO:")
print("""
| Coluna | Campo         | Tipo    | Descrição                    | Exemplo        |
|--------|---------------|---------|------------------------------|----------------|
| A      | Item          | VALOR   | Número sequencial            | 1, 2, 3...     |
| B      | Código        | VALOR   | Código do produto            | TRN10, DUN16   |
| C      | Produto       | FÓRMULA | VLOOKUP busca descrição      | (auto)         |
| D      | -             | -       | Não usado                    | -              |
| E      | -             | -       | Não usado                    | -              |
| F      | Embalagem     | VALOR   | Tipo de embalagem            | Bobina/Rolo    |
| G      | Lances        | VALOR   | Quantidade x metros          | 1x1000, 2x500  |
| H      | Quantidade    | VALOR   | Qtd total de unidades        | 1000           |
| I      | Valor Un.     | VALOR   | Preço unitário               | 3.74           |
| J      | Valor Total   | FÓRMULA | =I*H (calcula automatico)    | 3740.00        |
""")

print("\n📌 FÓRMULAS PRESERVAR EM VENDAS_PCP:")
print("""
| Célula  | Fórmula                                    |
|---------|-------------------------------------------|
| H6      | =J4+30                                    |
| H12     | =H8                                       |
| G15     | =C9                                       |
| C18:C32 | =IFERROR(VLOOKUP(B18:B32,N18:O198,2,0),"") |
| J18:J32 | =I18*H18 ... =I32*H32                     |
| I35     | =SUM(J18:J32)                             |
| E46     | =100%-E45                                 |
| I45     | =I35                                      |
""")

print("\n" + "-" * 80)
print("4️⃣ TABELA AUXILIAR DE PRODUTOS - Coluna N-O (linha 18 em diante)")
print("-" * 80)
print("Localização: N18:O198")
print("Coluna N = PRODUTO (código)")
print("Coluna O = DESCRIÇÃO")

# ===========================
# ANÁLISE PRODUÇÃO
# ===========================
ws2 = wb['PRODUÇÃO']

print("\n" + "=" * 120)
print("📋 PLANILHA PRODUÇÃO - ANÁLISE COMPLETA")
print("=" * 120)

print("\n" + "-" * 80)
print("1️⃣ CABEÇALHO (Dados vindos de VENDAS_PCP)")
print("-" * 80)

referencias_producao = [
    ("C4", "=VENDAS_PCP!C4", "Nº Orçamento"),
    ("E4", "=VENDAS_PCP!E4", "Revisão"),
    ("G4", "=VENDAS_PCP!G4", "Nº Pedido"),
    ("J4", "=VENDAS_PCP!J4", "Data Liberação"),
    ("C6", "=VENDAS_PCP!C6", "Vendedor"),
    ("H6", "=VENDAS_PCP!H6", "Prazo Entrega"),
    ("C7", "=VENDAS_PCP!C7", "Cliente"),
    ("C8", "=VENDAS_PCP!C8", "Contato"),
    ("H8", "=VENDAS_PCP!H8", "Telefone"),
    ("C9", "=VENDAS_PCP!C9", "Email"),
    ("J9", "=VENDAS_PCP!J9", "Tipo Frete"),
]

print(f"\n{'Célula':<8} | {'Fórmula':<25} | {'Descrição'}")
print("-" * 70)
for cell, formula, desc in referencias_producao:
    print(f"{cell:<8} | {formula:<25} | {desc}")

print("\n" + "-" * 80)
print("2️⃣ CABEÇALHO DE PRODUTOS - Linha 12")
print("-" * 80)

colunas_prod_producao = [
    ("A12", "-", "Não usado"),
    ("B12", "Cod.", "Código do produto"),
    ("C12", "Produto", "Descrição do produto"),
    ("D12", "-", "Não usado"),
    ("E12", "-", "Não usado"),
    ("F12", "Cod. Cores", "⭐ CÓDIGO DE CORES!"),
    ("G12", "-", "Não usado"),
    ("H12", "Embalagem:", "Tipo embalagem"),
    ("I12", "Lance(s)", "Formato lances"),
    ("J12", "Qtd.", "Quantidade"),
]

print(f"\n{'Célula':<8} | {'Cabeçalho':<15} | {'Descrição'}")
print("-" * 80)
for cell, header, desc in colunas_prod_producao:
    print(f"{cell:<8} | {header:<15} | {desc}")

print("\n" + "-" * 80)
print("3️⃣ ⭐⭐⭐ CÓDIGO DE CORES - DETALHAMENTO ⭐⭐⭐")
print("-" * 80)

print("""
📌 LOCALIZAÇÃO DO CÓDIGO DE CORES:
   - Planilha: PRODUÇÃO
   - Coluna de exibição: F (célula F13, F16, F19, etc)
   - Tabela de lookup: Coluna P (P18:P177)

📌 FÓRMULA USADA:
   F13: =IFERROR(VLOOKUP(B13,N18:P184,3,0),"")
   F16: =IFERROR(VLOOKUP(B16,N19:P188,3,0),"")
   ... e assim por diante

📌 ESTRUTURA DA TABELA DE CORES (Coluna N:P da PRODUÇÃO):
   | Coluna N | Coluna O    | Coluna P     |
   |----------|-------------|--------------|
   | PRODUTO  | DESCRIÇÃO   | Cod. Cores   |
   | DUN10    | DUPLA NET...| PT/NU        |
   | DUN16    | DUPLA NET...| PT/NU        |
   | TRN10    | TELA RECT...| PT/CZ/NU     |
   | TRN16    | TELA RECT...| PT/CZ/NU     |
   | TRI10    | TRIANGULAR..| PT/CZ/AZ     |
   | ...      | ...         | ...          |

📌 FORMATOS DE CÓDIGO DE CORES ENCONTRADOS:
""")

# Listar todos os códigos de cores únicos
codigos_cores = set()
for row in range(18, 178):
    cell = ws2.cell(row=row, column=16)  # Coluna P
    if cell.value and cell.value not in ['Cod. Cores', 'PRODUTO']:
        codigos_cores.add(str(cell.value))

print("   Códigos encontrados:")
for codigo in sorted(codigos_cores):
    print(f"   - {codigo}")

print("\n" + "-" * 80)
print("4️⃣ ÁREA DE PRODUTOS PRODUÇÃO - Estrutura de Blocos")
print("-" * 80)

print("""
📌 CADA PRODUTO OCUPA 3 LINHAS:

   LINHA PRINCIPAL (13, 16, 19, 22...):
   | Coluna | Campo        | Tipo    | Fórmula/Valor                |
   |--------|--------------|---------|------------------------------|
   | A      | Item         | VALOR   | 1, 2, 3...                   |
   | B      | Código       | FÓRMULA | =VENDAS_PCP!B18 (B19, B20...)|
   | C      | Produto      | FÓRMULA | VLOOKUP descrição            |
   | F      | Cod. Cores   | FÓRMULA | VLOOKUP para coluna P        |
   | H      | Embalagem    | FÓRMULA | =VENDAS_PCP!F18              |
   | I      | Lances       | FÓRMULA | =VENDAS_PCP!G18              |
   | J      | Quantidade   | FÓRMULA | =VENDAS_PCP!H18              |

   LINHA PESO/LOTE (14, 17, 20, 23...):
   | Coluna | Campo        | Tipo    | Descrição                    |
   |--------|--------------|---------|------------------------------|
   | A      | P. BRUTO     | VALOR   | Label "P. BRUTO"             |
   | D      | P.LIQUIDO    | VALOR   | Label "P.LIQUIDO"            |
   | F      | LOTE         | VALOR   | Label "LOTE"                 |
   | H      | (embalagem)  | FÓRMULA | =H13 (copia de cima)         |

   LINHA 3 (15, 18, 21, 24...):
   | (Geralmente vazia ou espaçamento)

📌 MAPEAMENTO PRODUTO VENDAS_PCP → PRODUÇÃO:
   
   | Produto | VENDAS_PCP | PRODUÇÃO (Principal) | PRODUÇÃO (Peso) |
   |---------|------------|---------------------|-----------------|
   | 1       | Linha 18   | Linha 13            | Linha 14        |
   | 2       | Linha 19   | Linha 16            | Linha 17        |
   | 3       | Linha 20   | Linha 19            | Linha 20        |
   | 4       | Linha 21   | Linha 22            | Linha 23        |
   | 5       | Linha 22   | Linha 25            | Linha 26        |
   | 6       | Linha 23   | Linha 28            | Linha 29        |
   | 7       | Linha 24   | Linha 31            | Linha 32        |
   | 8       | Linha 25   | Linha 34            | Linha 35        |
   | 9       | Linha 26   | Linha 37            | Linha 38        |
   | 10      | Linha 27   | Linha 40            | Linha 41        |
   | 11      | Linha 28   | Linha 43            | Linha 44        |
   | 12      | Linha 29   | Linha 46            | Linha 47        |
   | 13      | Linha 30   | Linha 49            | Linha 50        |
   | 14      | Linha 31   | Linha 52            | Linha 53        |
   | 15      | Linha 32   | Linha 55            | Linha 56        |
""")

print("\n" + "=" * 120)
print("📊 MAPEAMENTO COMPLETO - TABELA FINAL")
print("=" * 120)

print("""
┌─────────────┬─────────┬─────────────────────────────┬──────────┬────────────────────┐
│ PLANILHA    │ CÉLULA  │ DESCRIÇÃO                   │ TIPO     │ EXEMPLO            │
├─────────────┼─────────┼─────────────────────────────┼──────────┼────────────────────┤
│             │         │ === CABEÇALHO ===           │          │                    │
│ VENDAS_PCP  │ C4      │ Número do Orçamento         │ VALOR    │ 352                │
│ VENDAS_PCP  │ E4      │ Revisão                     │ VALOR    │ (vazio ou número)  │
│ VENDAS_PCP  │ G4      │ Número do Pedido            │ VALOR    │ 202500083          │
│ VENDAS_PCP  │ J4      │ Data de Liberação           │ VALOR    │ 2025-08-19         │
│ VENDAS_PCP  │ C6      │ Vendedor                    │ VALOR    │ Marcia Scarcella   │
│ VENDAS_PCP  │ H6      │ Prazo de Entrega            │ FÓRMULA  │ =J4+30             │
│ VENDAS_PCP  │ C7      │ Cliente                     │ VALOR    │ CONSTRULAR         │
│ VENDAS_PCP  │ C8      │ Contato                     │ VALOR    │ Rodrigo            │
│ VENDAS_PCP  │ H8      │ Telefone                    │ VALOR    │ 94984306216        │
│ VENDAS_PCP  │ C9      │ Email                       │ VALOR    │ email@gmail.com    │
│ VENDAS_PCP  │ J9      │ Tipo de Frete               │ VALOR    │ FOB                │
│ VENDAS_PCP  │ C12     │ Transportadora              │ VALOR    │ (nome)             │
│ VENDAS_PCP  │ C13     │ CEP                         │ VALOR    │ 68560-000          │
│ VENDAS_PCP  │ F13     │ Endereço Completo           │ VALOR    │ Rua/Av...          │
│ VENDAS_PCP  │ C15     │ CPF/CNPJ                    │ VALOR    │ 36408556000169     │
├─────────────┼─────────┼─────────────────────────────┼──────────┼────────────────────┤
│             │         │ === PRODUTOS (18-32) ===    │          │                    │
│ VENDAS_PCP  │ A18-A32 │ Número do Item              │ VALOR    │ 1, 2, 3...         │
│ VENDAS_PCP  │ B18-B32 │ Código do Produto           │ VALOR    │ TRN10, DUN16       │
│ VENDAS_PCP  │ C18-C32 │ Descrição (VLOOKUP)         │ FÓRMULA  │ TELA RECT NET...   │
│ VENDAS_PCP  │ F18-F32 │ Embalagem                   │ VALOR    │ Bobina, Rolo       │
│ VENDAS_PCP  │ G18-G32 │ Lances                      │ VALOR    │ 1x1000, 2x500      │
│ VENDAS_PCP  │ H18-H32 │ Quantidade                  │ VALOR    │ 1000               │
│ VENDAS_PCP  │ I18-I32 │ Valor Unitário              │ VALOR    │ 3.74               │
│ VENDAS_PCP  │ J18-J32 │ Valor Total (auto)          │ FÓRMULA  │ =I18*H18           │
├─────────────┼─────────┼─────────────────────────────┼──────────┼────────────────────┤
│             │         │ === TOTAIS ===              │          │                    │
│ VENDAS_PCP  │ I35     │ Total do Pedido             │ FÓRMULA  │ =SUM(J18:J32)      │
│ VENDAS_PCP  │ E45     │ % Pagamento                 │ VALOR    │ 1 (100%)           │
│ VENDAS_PCP  │ F45     │ Método Pagamento            │ VALOR    │ FATURAMENTO        │
│ VENDAS_PCP  │ I45     │ Valor Total                 │ FÓRMULA  │ =I35               │
├─────────────┼─────────┼─────────────────────────────┼──────────┼────────────────────┤
│             │         │ === PRODUÇÃO ===            │          │                    │
│ PRODUÇÃO    │ C4      │ Orçamento                   │ FÓRMULA  │ =VENDAS_PCP!C4     │
│ PRODUÇÃO    │ E4      │ Revisão                     │ FÓRMULA  │ =VENDAS_PCP!E4     │
│ PRODUÇÃO    │ G4      │ Pedido                      │ FÓRMULA  │ =VENDAS_PCP!G4     │
│ PRODUÇÃO    │ J4      │ Data Liberação              │ FÓRMULA  │ =VENDAS_PCP!J4     │
│ PRODUÇÃO    │ C6      │ Vendedor                    │ FÓRMULA  │ =VENDAS_PCP!C6     │
│ PRODUÇÃO    │ H6      │ Prazo Entrega               │ FÓRMULA  │ =VENDAS_PCP!H6     │
│ PRODUÇÃO    │ C7      │ Cliente                     │ FÓRMULA  │ =VENDAS_PCP!C7     │
│ PRODUÇÃO    │ C8      │ Contato                     │ FÓRMULA  │ =VENDAS_PCP!C8     │
│ PRODUÇÃO    │ H8      │ Telefone                    │ FÓRMULA  │ =VENDAS_PCP!H8     │
│ PRODUÇÃO    │ C9      │ Email                       │ FÓRMULA  │ =VENDAS_PCP!C9     │
│ PRODUÇÃO    │ J9      │ Frete                       │ FÓRMULA  │ =VENDAS_PCP!J9     │
├─────────────┼─────────┼─────────────────────────────┼──────────┼────────────────────┤
│             │         │ === PRODUTOS PRODUÇÃO ===   │          │                    │
│ PRODUÇÃO    │ B13     │ Código Produto 1            │ FÓRMULA  │ =VENDAS_PCP!B18    │
│ PRODUÇÃO    │ C13     │ Descrição 1                 │ FÓRMULA  │ VLOOKUP            │
│ PRODUÇÃO    │ F13     │ ⭐ CÓDIGO CORES 1           │ FÓRMULA  │ VLOOKUP P          │
│ PRODUÇÃO    │ H13     │ Embalagem 1                 │ FÓRMULA  │ =VENDAS_PCP!F18    │
│ PRODUÇÃO    │ I13     │ Lances 1                    │ FÓRMULA  │ =VENDAS_PCP!G18    │
│ PRODUÇÃO    │ J13     │ Quantidade 1                │ FÓRMULA  │ =VENDAS_PCP!H18    │
├─────────────┼─────────┼─────────────────────────────┼──────────┼────────────────────┤
│             │         │ === TABELA CORES ===        │          │                    │
│ PRODUÇÃO    │ N18     │ Código Produto (lookup)     │ VALOR    │ DUN10, TRN10...    │
│ PRODUÇÃO    │ O18     │ Descrição (lookup)          │ VALOR    │ DUPLA NET 10mm...  │
│ PRODUÇÃO    │ P18     │ Código de Cores (lookup)    │ VALOR    │ PT/NU, PT/CZ/NU    │
└─────────────┴─────────┴─────────────────────────────┴──────────┴────────────────────┘
""")

print("\n" + "=" * 120)
print("🎨 CÓDIGOS DE CORES - LEGENDA")
print("=" * 120)

print("""
Abreviações encontradas:
  PT = Preto
  CZ = Cinza
  NU = Natural/Nude
  AZ = Azul
  VM = Vermelho
  VD = Verde
  AM = Amarelo
  BC = Branco
  LR = Laranja
  MR = Marrom

Formatos comuns:
  - PT/NU        → 2 cores (Preto/Natural)
  - PT/CZ/NU     → 3 cores (Preto/Cinza/Natural)
  - PT/CZ/AZ     → 3 cores (Preto/Cinza/Azul)
  - PT/CZ/VM/NU  → 4 cores
  - PT/CZ/VM/AZ  → 4 cores
  - AM/VD/VM/PT  → 4 cores (sinalização)
  - AM/VD/VM/AZ/PT/BC/MR    → 7 cores
  - AM/VD/VM/AZ/PT/BC/LR/MR → 8 cores
""")

wb.close()
print("\n✅ Análise completa!")
