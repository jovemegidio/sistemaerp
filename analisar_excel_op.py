"""
Script para analisar o arquivo Excel de Ordem de Produção Aluforce
"""
import openpyxl
from openpyxl.utils import get_column_letter
import json

# Caminho do arquivo
arquivo = r"c:\Users\egidio\Documents\Sistema - ALUFORCE - V.2\modules\PCP\Ordem de Produção Aluforce - Copia.xlsx"

print("=" * 80)
print("ANÁLISE DO ARQUIVO EXCEL - ORDEM DE PRODUÇÃO ALUFORCE")
print("=" * 80)

# Carregar workbook
wb = openpyxl.load_workbook(arquivo, data_only=False)

# 1. Listar todas as planilhas
print("\n1. PLANILHAS EXISTENTES NO ARQUIVO:")
print("-" * 40)
for i, sheet_name in enumerate(wb.sheetnames, 1):
    ws = wb[sheet_name]
    print(f"   {i}. {sheet_name} (Linhas: {ws.max_row}, Colunas: {ws.max_column})")

# 2. Analisar a planilha VENDAS_PCP
print("\n2. ANÁLISE DA PLANILHA 'VENDAS_PCP':")
print("-" * 40)

if 'VENDAS_PCP' in wb.sheetnames:
    ws_vendas = wb['VENDAS_PCP']
    
    print(f"   Dimensões: {ws_vendas.max_row} linhas x {ws_vendas.max_column} colunas")
    
    # Verificar cabeçalhos importantes
    print("\n   CABEÇALHOS (linha 17 e ao redor):")
    for row in range(15, 20):
        row_data = []
        for col in range(1, 8):
            cell = ws_vendas.cell(row=row, column=col)
            if cell.value:
                row_data.append(f"{get_column_letter(col)}{row}={cell.value}")
        if row_data:
            print(f"   Linha {row}: {', '.join(row_data)}")
    
    # 3. Analisar células C18:E32
    print("\n3. CONTEÚDO DAS CÉLULAS C18:E32:")
    print("-" * 40)
    
    for row in range(18, 33):
        for col in ['C', 'D', 'E']:
            col_idx = ord(col) - ord('A') + 1
            cell = ws_vendas.cell(row=row, column=col_idx)
            
            # Verificar se tem fórmula
            if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                print(f"   Célula {col}{row}: FÓRMULA = {cell.value}")
            elif cell.value:
                print(f"   Célula {col}{row}: VALOR = {cell.value}")
            # Mostrar também se vazio mas queremos ver todas
    
    # 4. Buscar por todas as fórmulas VLOOKUP no arquivo
    print("\n4. TODAS AS FÓRMULAS VLOOKUP NO ARQUIVO:")
    print("-" * 40)
    
    vlookup_found = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in range(1, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                if cell.value and isinstance(cell.value, str):
                    cell_str = str(cell.value).upper()
                    if 'VLOOKUP' in cell_str or 'PROCV' in cell_str:
                        vlookup_found.append({
                            'planilha': sheet_name,
                            'celula': f"{get_column_letter(col)}{row}",
                            'formula': cell.value
                        })
                        print(f"   [{sheet_name}] {get_column_letter(col)}{row}: {cell.value}")
    
    if not vlookup_found:
        print("   Nenhuma fórmula VLOOKUP/PROCV encontrada no arquivo.")

else:
    print("   ERRO: Planilha 'VENDAS_PCP' não encontrada!")

# 5. Verificar se existe planilha de produtos
print("\n5. BUSCA POR PLANILHA/TABELA DE PRODUTOS:")
print("-" * 40)

planilhas_produtos = []
for sheet_name in wb.sheetnames:
    name_lower = sheet_name.lower()
    if 'produto' in name_lower or 'item' in name_lower or 'catalogo' in name_lower or 'lista' in name_lower:
        planilhas_produtos.append(sheet_name)
        print(f"   Encontrada: {sheet_name}")

if not planilhas_produtos:
    print("   Nenhuma planilha específica de produtos encontrada.")
    print("   Verificando se há dados de produtos em outras planilhas...")

# 6. Analisar estrutura de cada planilha
print("\n6. ESTRUTURA DETALHADA DE CADA PLANILHA:")
print("-" * 40)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n   === {sheet_name} ===")
    
    # Mostrar primeiras linhas para entender estrutura
    print("   Primeiras 10 linhas (colunas A-H):")
    for row in range(1, min(11, ws.max_row + 1)):
        row_data = []
        for col in range(1, min(9, ws.max_column + 1)):
            cell = ws.cell(row=row, column=col)
            val = cell.value
            if val:
                if isinstance(val, str) and len(val) > 30:
                    val = val[:27] + "..."
                row_data.append(f"{get_column_letter(col)}:{val}")
        if row_data:
            print(f"      L{row}: {' | '.join(row_data)}")

# 7. Verificar Named Ranges (intervalos nomeados)
print("\n7. INTERVALOS NOMEADOS (NAMED RANGES):")
print("-" * 40)

if wb.defined_names:
    for name in wb.defined_names.definedName:
        print(f"   Nome: {name.name}")
        print(f"   Referência: {name.attr_text}")
        print()
else:
    print("   Nenhum intervalo nomeado encontrado.")

# 8. Buscar todas as fórmulas no range C18:E32
print("\n8. ANÁLISE DETALHADA C18:E32 (incluindo células vazias):")
print("-" * 40)

if 'VENDAS_PCP' in wb.sheetnames:
    ws_vendas = wb['VENDAS_PCP']
    
    for row in range(18, 33):
        print(f"\n   Linha {row}:")
        for col_letter in ['C', 'D', 'E']:
            col_idx = ord(col_letter) - ord('A') + 1
            cell = ws_vendas.cell(row=row, column=col_idx)
            
            valor = cell.value
            tipo = type(valor).__name__
            
            tem_formula = False
            if valor and isinstance(valor, str) and valor.startswith('='):
                tem_formula = True
            
            print(f"      {col_letter}{row}: Valor='{valor}' | Tipo={tipo} | Formula={tem_formula}")

# 9. Verificar tabelas do Excel
print("\n9. TABELAS DO EXCEL (ListObjects):")
print("-" * 40)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    if hasattr(ws, 'tables') and ws.tables:
        for table_name, table in ws.tables.items():
            print(f"   Planilha: {sheet_name}")
            print(f"   Nome da tabela: {table_name}")
            print(f"   Range: {table.ref}")
            print()

print("\n" + "=" * 80)
print("FIM DA ANÁLISE")
print("=" * 80)

wb.close()
