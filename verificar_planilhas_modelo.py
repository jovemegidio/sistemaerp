import openpyxl

wb = openpyxl.load_workbook('modules/PCP/Ordem de Produção Aluforce - Copia.xlsx')

print('PLANILHAS NO ARQUIVO:')
print('='*60)
for i, sheet_name in enumerate(wb.sheetnames, 1):
    ws = wb[sheet_name]
    print(f'{i}. {sheet_name} ({ws.max_row} linhas x {ws.max_column} colunas)')

# Verificar se existe VENDAS_PCP
if 'VENDAS_PCP' in wb.sheetnames:
    print('\n[OK] Planilha VENDAS_PCP encontrada!')
    ws_vendas = wb['VENDAS_PCP']
    print(f'Dimensoes: {ws_vendas.max_row} linhas x {ws_vendas.max_column} colunas')
    
    print('\nPrimeiras celulas com dados:')
    for row in range(1, min(30, ws_vendas.max_row + 1)):
        valores_linha = []
        for col in range(1, min(15, ws_vendas.max_column + 1)):
            cell = ws_vendas.cell(row, col)
            if cell.value:
                col_letter = openpyxl.utils.get_column_letter(col)
                valores_linha.append(f'{col_letter}{row}={str(cell.value)[:30]}')
        if valores_linha:
            print(f'Linha {row}: {" | ".join(valores_linha[:5])}')
else:
    print('\n[AVISO] Planilha VENDAS_PCP NAO encontrada!')
    print('As formulas =VENDAS_PCP!... nao funcionarao sem essa planilha')

wb.close()
