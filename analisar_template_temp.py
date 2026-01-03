import openpyxl

wb = openpyxl.load_workbook(r'C:\Users\egidio\Documents\Sistema - ALUFORCE - V.2\modules\PCP\ORDEM_COMPLETA_TESTE.xlsx')
ws = wb['VENDAS_PCP']

print('=== MAPEAMENTO CORRETO DAS CELULAS ===\n')

# Linhas 4-15 detalhadas
for i in range(4, 16):
    cells = []
    for col in 'ABCDEFGHIJ':
        val = ws[f'{col}{i}'].value
        if val:
            val_str = str(val)[:35]
            cells.append(f'{col}={val_str}')
    if cells:
        print(f'Linha {i}: {cells}')

print('\n=== PRODUTOS (Linhas 17-20) ===\n')
for i in range(17, 21):
    cells = []
    for col in 'ABCDEFGHIJ':
        val = ws[f'{col}{i}'].value
        if val:
            val_str = str(val)[:25]
            cells.append(f'{col}={val_str}')
    if cells:
        print(f'Linha {i}: {cells}')
